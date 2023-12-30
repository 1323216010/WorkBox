import os
from configparser import ConfigParser
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
import xlrd

#控制台显示excel和csv文件读取过程
def read_file_with_progress(filename, all_sheets=False):
    _, ext = os.path.splitext(filename)  # 获取文件扩展名

    if ext in ['.xlsx', '.xls']:
        # 根据文件扩展名选择适当的库
        workbook_loader = load_workbook if ext == '.xlsx' else xlrd.open_workbook
        workbook = workbook_loader(filename, read_only=True)

        if all_sheets:
            sheets_data = {}  # 用于存储所有工作表的数据
            sheet_names = workbook.sheetnames if ext == '.xlsx' else workbook.sheet_names()

            for sheet_name in sheet_names:
                sheet = workbook[sheet_name] if ext == '.xlsx' else workbook.sheet_by_name(sheet_name)
                total_rows = sheet.max_row if ext == '.xlsx' else sheet.nrows

                data = []
                row_iterator = sheet.iter_rows() if ext == '.xlsx' else range(total_rows)
                for row in tqdm(row_iterator, total=total_rows, desc=f'Reading {sheet_name}'):
                    row_data = [cell.value for cell in row] if ext == '.xlsx' else sheet.row_values(row)
                    data.append(row_data)

                df = pd.DataFrame(data[1:], columns=data[0])
                sheets_data[sheet_name] = df  # 将数据框添加到字典中

            return sheets_data  # 返回包含所有工作表数据的字典
        else:
            sheet = workbook.active if ext == '.xlsx' else workbook.sheet_by_index(0)
            total_rows = sheet.max_row if ext == '.xlsx' else sheet.nrows

            data = []
            row_iterator = sheet.iter_rows() if ext == '.xlsx' else range(total_rows)
            for row in tqdm(row_iterator, total=total_rows):
                row_data = [cell.value for cell in row] if ext == '.xlsx' else sheet.row_values(row)
                data.append(row_data)

            df = pd.DataFrame(data[1:], columns=data[0])

            return {workbook.sheetnames[0]: df}  # 返回第一张工作表的数据框
    elif ext == '.csv':
        # 使用pandas的read_csv函数，但是分块读取
        chunk_size = 5000  # 可以根据需要调整块大小
        chunks = pd.read_csv(filename, chunksize=chunk_size)

        # 获取文件的总行数，以便我们可以显示进度
        total_rows = sum(1 for row in open(filename))

        data = []
        for chunk in tqdm(chunks, total=total_rows // chunk_size):
            data.append(chunk)

        # 合并所有块为一个DataFrame
        df = pd.concat(data, axis=0)
        return {'sheet1': df}
    else:
        raise ValueError(f"Unsupported file extension: {ext}")

def getConfig():
    dict1 = {}
    config = ConfigParser()
    fp_dir = os.getcwd()
    path = os.path.join(fp_dir, "config.ini")
    config.read(path)

    # data
    dict1["pathdata"] = config['data']['pathdata']
    dict1["corr"] = config['data']['corr']
    dict1["Serial_Number"] = config['data']['Serial_Number']

    # spec
    dict1["pathspec"] = config['spec']['pathspec']
    dict1["USL"] = config['spec']['USL']
    dict1["LSL"] = config['spec']['LSL']
    dict1["uncorrs"] = [dict1["corr"], dict1["Serial_Number"]]


    # font
    dict1["font"] = "yan"
    try:
        dict1["font_path"] = config['font']['font_path']
    except:
        dict1["font_path"] = "./ttf/DejaVuSans.ttf"
    try:
        dict1["font_size"] = int(config['font']['font_size'])
    except:
        dict1["font_size"] = 17
    try:
        dict1["font_high"] = int(config['font']['high'])
    except:
        dict1["font_high"] = 0

    # rule
    try:
        dict1["dot_size"] = int(config['rule']['dot_size'])
    except:
        dict1["dot_size"] = 20
    try:
        dict1["r2_position"] = config['rule']['r2_position']
    except:
        dict1["r2_position"] = "outside"
    try:
        dict1["depend_spec"] = config['rule']['depend_spec']
    except:
        dict1["depend_spec"] = "off"
    try:
        dict1["axis_range"] = config['rule']['axis_range']
    except:
        dict1["axis_range"] = "off"
    try:
        dict1["axis_x1"] = config['rule']['axis_x1']
    except:
        dict1["axis_x1"] = "x1"
    try:
        dict1["axis_x2"] = config['rule']['axis_x2']
    except:
        dict1["axis_x2"] = "x2"
    try:
        dict1["axis_y1"] = config['rule']['axis_y1']
    except:
        dict1["axis_y1"] = "y1"
    try:
        dict1["axis_y2"] = config['rule']['axis_y2']
    except:
        dict1["axis_y2"] = "y2"

    #box
    try:
        dict1["box_picture"] = config['box']['box_picture']
    except:
        dict1["box_picture"] = "1"
    try:
        dict1["box_img_w"] = int(config['box']['box_img_w'])
    except:
        dict1["box_img_w"] = 1280
    try:
        dict1["box_img_x"] = int(config['box']['box_img_x'])
    except:
        dict1["box_img_x"] = 320
    try:
        dict1["box_img_dpi"] = int(config['box']['box_img_dpi'])
    except:
        dict1["box_img_dpi"] = 100
    try:
        dict1["box_font_size"] = int(config['box']['box_font_size'])
    except:
        dict1["box_font_size"] = 8
    try:
        dict1["box_decimal_format"] = int(config['box']['decimal_format'])
    except:
        dict1["box_decimal_format"] = 2

    # bar
    try:
        dict1["hist_quantity"] = int(config['hist']['quantity'])
    except:
        dict1["hist_quantity"] = 20
    try:
        dict1["hist_decimal_format"] = int(config['hist']['decimal_format'])
    except:
        dict1["hist_decimal_format"] = 2
    try:
        dict1["hist_alpha"] = float(config['hist']['alpha'])
    except:
        dict1["hist_alpha"] = 0.5
    try:
        dict1["hist_label_position_auto"] = config['hist']['label_position_auto']
    except:
        dict1["hist_label_position_auto"] = "off"
    try:
        dict1["hist_label_size"] = int(config['hist']['label_size'])
    except:
        dict1["hist_label_size"] = 10
    try:
        dict1["hist_grid"] = config['hist']['grid']
    except:
        dict1["hist_grid"] = "off"
    try:
        dict1["hist_grid_linewidth"] = float(config['hist']['grid_linewidth'])
    except:
        dict1["hist_grid_linewidth"] = 0.5

    # xy
    try:
        dict1["xy_decimal_format"] = int(config['xy']['decimal_format'])
    except:
        dict1["xy_decimal_format"] = 2
    try:
        dict1["xy_grid"] = config['xy']['grid']
    except:
        dict1["xy_grid"] = "off"
    try:
        dict1["xy_grid_linewidth"] = float(config['xy']['grid_linewidth'])
    except:
        dict1["xy_grid_linewidth"] = 0.5
    try:
        dict1["xy_spec_set_y"] = config['xy']['spec_set_y']
    except:
        dict1["xy_spec_set_y"] = "off"
    try:
        dict1["xy_spec_y_usl"] = config['xy']['spec_y_usl']
    except:
        dict1["xy_spec_y_usl"] = "y_usl"
    try:
        dict1["xy_spec_y_lsl"] = config['xy']['spec_y_lsl']
    except:
        dict1["xy_spec_y_lsl"] = "y_lsl"

    # type4
    try:
        dict1["type4_xlabel"] = int(config['type4']['xlabel'])
    except:
        dict1["type4_xlabel"] = 2
    try:
        dict1["type4_xlabel_size"] = int(config['type4']['xlabel_size'])
    except:
        dict1["type4_xlabel_size"] = 13
    try:
        dict1["type4_xlabel1_position"] = int(config['type4']['xlabel1_position'])
    except:
        dict1["type4_xlabel1_position"] = 14
    try:
        dict1["type4_xlabel2_position"] = int(config['type4']['xlabel2_position'])
    except:
        dict1["type4_xlabel2_position"] = -5

    return dict1

def print_version():
    print("DrawToPDF version is 20231220b")
    print("=========================================")
    print("Update Log")
    print("=========================================")
    print("[2023-12-20]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add the box_font_size and decimal_format configuration options for box")
    print("    - Add standard deviation for the box plot")
    print("Fixes:")
    print("    - Fix the issue of list order disruption in ‘depend_spec=on’ mode by employing list comprehension for filtering operations")
    print("[2023-12-06]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add configuration options for Y-axis specifications")
    print("[2023-11-30]")
    print("-----------------------------------------")
    print("Features:")
    print("    - For type4, add a configuration for the number of X-axis position labels and align colons when the number is 2")
    print("[2023-11-08]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add hist and xy decimal place reservation configurations")
    print("    - Add the hist tag location configuration")
    print("[2023-11-07]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add the hist plot")
    print("[2023-10-19]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add configuration for the number of images, horizontal position, width, and resolution for mode 3")
    print("Fixes:")
    print("    - Fixed an issue where the configuration file could not be modified before input")
    print("[2023-10-14]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add the mean label and position adaptation to the box plot")
    print("[2023-10-13]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add a configuration switch whether to customize the axes")
    print("    - Add the x-axis and y-axis configurations")
    print("[2023-10-12]")
    print("-----------------------------------------")
    print("Features:")
    print("    - Add a configuration for the vertical position of the title")
    print("    - Add r2_position configuration")
    print("    - Add depend_spec configuration")
    print("    - Modify the axis selection rules")
    print("    - The box diagram no longer limits the number of items")
    print("Fixes:")
    print("    - Fixed an issue with the accuracy of October version R2_score")
    print()
    print("For any questions or concerns, please contact:")
    print("Pengcheng.yan@cowellchina.com")
    print("=========================================")

    print("Please select a type of pdf:")
    print("Type 0: xy and scatter plots")
    print("Type 1: only scatter plots")
    print("Type 2: only xy plots")
    print("Type 3: only box plots")
    print("Type 4: xy and hist plots")
    # print("Type 4: only curve plots")